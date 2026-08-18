# -*- coding: utf-8 -*-
"""아직 못 채운 정답을 손으로 넣기 위한 준비.

  ① 읽어야 할 해설 PDF를 «데이터베이스\\정답 입력할 파일» 한 폴더로 복사한다 (원본은 그대로 둔다)
  ② «정답 입력표.xlsx» 를 만든다 — 한 줄에 한 회차, 가로로 1~30번 칸
       · 이미 넣은 정답은 회색으로 채워 두고, 빈칸만 치시면 된다
       · 파일이 아직 없는 회차는 «구해야 할 목록» 시트에 따로 적는다
  ③ 다 채우신 엑셀은 read_answer_form.py 가 읽어 data/ans.js 에 합친다
"""
import sys, os, re, json, glob, shutil, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_answer as BA
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.dirname(HERE)
DBROOT = os.path.abspath(os.path.join(APP, '..', '데이터베이스'))
DB = os.path.join(DBROOT, '해설파일(정리 전)')
COPYTO = os.path.join(DBROOT, '정답 입력할 파일')
XLSX = os.path.join(DBROOT, '정답 입력표.xlsx')

FONT = 'Arial'
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
DONE_FILL = PatternFill('solid', fgColor='EDEDED')     # 이미 넣은 정답
TODO_FILL = PatternFill('solid', fgColor='FFF7E0')     # 치셔야 할 빈칸
NA_FILL = PatternFill('solid', fgColor='F7F7F7')       # 그 시험에 없는 번호
THIN = Side(style='thin', color='C9CDD3')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load():
    src = open(os.path.join(APP, 'data', 'index.js'), encoding='utf-8').read()
    Q = json.loads(src[src.index('{'):src.rindex('}') + 1].rstrip(';'))
    ans = json.loads(open(os.path.join(APP, 'data', 'ans.js'), encoding='utf-8')
                     .read().split('=', 1)[1].rstrip(';\n'))
    return Q, ans


def exam_files(Q):
    """시험 이름 → 그 정답이 들어 있는 해설 파일 (build_answer 와 같은 규칙)"""
    exams = Q['exams']
    by_key = collections.defaultdict(list)
    for i, e in enumerate(exams):
        by_key[(e['g'], e['y'], e['r'])].append(i)

    files = sorted(glob.glob(os.path.join(DB, '*', '*.pdf'))) + \
            sorted(glob.glob(os.path.join(DB, '*', '해설', '*.pdf'))) + \
            sorted(glob.glob(os.path.join(DB, '*', '문항', '*.pdf')))
    files = [f for f in files if '_지울것' not in f and '_확인필요' not in f]

    out = {}
    for f in files:
        rel = os.path.relpath(f, DB)
        key = BA.parse_name(rel)
        if not key:
            continue
        g, y, r, hyung = key
        # 파일 이름에 과목이 박혀 있으면 (2021년 7월처럼) 그 과목에만 맡긴다
        base = os.path.basename(rel)
        only = None
        for k, v in (('확통', ('확통', '확률과 통계')), ('미적', ('미적', '미적분')),
                     ('기하', ('기하',))):
            if k in base:
                only = v
                break
        for i in by_key.get((g, y, r), []):
            s = exams[i]['s']
            if hyung and s not in ('공통', '', hyung):
                continue
            if only and s not in only:
                continue
            out.setdefault(exams[i]['n'], rel)
    return out


def safe(s):
    return re.sub(r'[\\/:*?"<>|]', '_', s)


def main():
    Q, ANS = load()
    exams = Q['exams']
    qno = collections.defaultdict(set)
    for r in Q['items']:
        qno[exams[r[0]]['n']].add(r[1])
    src = exam_files(Q)

    rows_todo, rows_need = [], []
    for e in exams:
        n = e['n']
        need = sorted(qno[n])
        have = {int(k): v for k, v in ANS.get(n, {}).items()}
        miss = [q for q in need if q not in have]
        if not miss:
            continue
        rec = {'e': e, 'need': need, 'have': have, 'miss': miss, 'file': src.get(n)}
        (rows_todo if rec['file'] else rows_need).append(rec)

    # ── ① 파일 모으기 ────────────────────────────────
    os.makedirs(COPYTO, exist_ok=True)
    for old in glob.glob(os.path.join(COPYTO, '*.pdf')):
        os.remove(old)
    copied = {}
    for rec in rows_todo:
        rel = rec['file']
        if rel in copied:
            continue
        base = os.path.basename(rel)
        folder = rel.split(os.sep)[0]
        new = safe(folder.replace('전국연합_', '').replace('(3월 제외)', '')
                   .replace('_3월(중학 범위)', '').replace('_3월(고1 범위)', '')) + '_' + base
        shutil.copy2(os.path.join(DB, rel), os.path.join(COPYTO, new))
        copied[rel] = new
    for rec in rows_todo:
        rec['copy'] = copied[rec['file']]

    # ── ② 엑셀 ──────────────────────────────────────
    wb = Workbook()

    ws = wb.active
    ws.title = '정답 입력'
    head = ['시험', '학년', '연도', '시행', '과목', '읽을 파일', '빈칸 수'] + \
           [str(i) for i in range(1, 31)]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BOX
    ws.freeze_panes = 'H2'

    r = 2
    for rec in sorted(rows_todo, key=lambda x: (x['e']['g'], x['e']['y'], x['e']['r'], x['e']['s'])):
        e, have, need = rec['e'], rec['have'], rec['need']
        ws.cell(r, 1, e['n']); ws.cell(r, 2, e['g']); ws.cell(r, 3, e['y'])
        ws.cell(r, 4, e['r']); ws.cell(r, 5, e['s'] or '단일')
        ws.cell(r, 6, rec['copy'])
        # «-» 칸에는 글자가 들어 있어 빈칸으로 세지 않는다. 그래서 COUNTBLANK 하나면 된다.
        ws.cell(r, 7, '=COUNTBLANK(H%d:AK%d)' % (r, r))
        for q in range(1, 31):
            c = ws.cell(r, 7 + q)
            c.border = BOX
            c.alignment = Alignment(horizontal='center')
            c.font = Font(name=FONT, size=10)
            if q not in need:
                c.value = '-'; c.fill = NA_FILL
                c.font = Font(name=FONT, size=10, color='BFBFBF')
            elif q in have:
                c.value = have[q]; c.fill = DONE_FILL
                c.font = Font(name=FONT, size=10, color='808080')
            else:
                c.fill = TODO_FILL
        for c in range(1, 8):
            ws.cell(r, c).font = Font(name=FONT, size=10)
            ws.cell(r, c).border = BOX
        r += 1

    ws.column_dimensions['A'].width = 26
    for col, w in zip('BCDE', (7, 7, 7, 11)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions['F'].width = 34
    ws.column_dimensions['G'].width = 8
    for q in range(1, 31):
        ws.column_dimensions[get_column_letter(7 + q)].width = 4.2

    # ── 구해야 할 목록 ──────────────────────────────
    w2 = wb.create_sheet('구해야 할 목록')
    h2 = ['시험', '학년', '연도', '시행', '과목', '문항 수', '비고']
    w2.append(h2)
    for c in range(1, len(h2) + 1):
        cell = w2.cell(1, c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = BOX
    r = 2
    for rec in sorted(rows_need, key=lambda x: (x['e']['g'], x['e']['y'], x['e']['r'], x['e']['s'])):
        e = rec['e']
        w2.cell(r, 1, e['n']); w2.cell(r, 2, e['g']); w2.cell(r, 3, e['y'])
        w2.cell(r, 4, e['r']); w2.cell(r, 5, e['s'] or '단일')
        w2.cell(r, 6, len(rec['need']))
        w2.cell(r, 7, '평가원 정답표 (수능·모평)' if e['g'] == '수능·모평' else '해설지나 정답지를 아직 못 구함')
        for c in range(1, 8):
            w2.cell(r, c).font = Font(name=FONT, size=10)
            w2.cell(r, c).border = BOX
        r += 1
    w2.column_dimensions['A'].width = 26
    for col, wd in zip('BCDEF', (7, 7, 7, 11, 8)):
        w2.column_dimensions[col].width = wd
    w2.column_dimensions['G'].width = 30
    w2.freeze_panes = 'A2'

    # ── 읽는 법 ────────────────────────────────────
    w3 = wb.create_sheet('쓰는 법', 0)
    guide = [
        ['정답 입력표 — 쓰는 법', ''],
        ['', ''],
        ['치실 곳', '«정답 입력» 시트의 노란 칸(H~AK열)에만 숫자를 넣어 주세요.'],
        ['', '객관식은 1~5, 단답형은 답 그대로(예: 296) 적으시면 됩니다.'],
        ['회색 칸', '이미 넣어 둔 정답입니다. 고치지 않으셔도 됩니다.'],
        ['«-» 칸', '그 시험에 없는 번호입니다. 비워 두세요.'],
        ['읽을 파일', '«데이터베이스\\정답 입력할 파일» 폴더에 같은 이름의 PDF가 있습니다.'],
        ['', '그 PDF 첫 쪽의 «빠른 정답» 표를 보고 옮겨 적으시면 됩니다.'],
        ['빈칸 수', '아직 안 채운 칸이 몇 개인지 저절로 세어 줍니다. 0이 되면 그 줄은 끝난 것입니다.'],
        ['', ''],
        ['보기 (첫 줄에 이렇게 들어갑니다)', ''],
        ['시험', '2018년 3월 고1'],
        ['읽을 파일', '고1_1803_고1_해설.pdf'],
        ['22번 칸', '48'],
        ['23번 칸', '22'],
        ['', ''],
        ['다 채우신 뒤', '저에게 «정답 입력표 다 채웠어» 라고 알려 주시면 검색기에 넣겠습니다.'],
        ['', '한 줄만 채우셔도 됩니다. 채운 만큼만 넣습니다.'],
        ['', ''],
        ['구해야 할 목록', '해설지·정답지가 아직 없는 회차입니다. 파일을 구해 오시면 그때 넣습니다.'],
    ]
    for row in guide:
        w3.append(row)
    w3['A1'].font = Font(name=FONT, bold=True, size=14)
    w3['A11'].font = Font(name=FONT, bold=True, size=11)
    for i in range(3, len(guide) + 1):
        w3.cell(i, 1).font = Font(name=FONT, bold=True, size=10)
        w3.cell(i, 2).font = Font(name=FONT, size=10)
    w3.column_dimensions['A'].width = 30
    w3.column_dimensions['B'].width = 78

    wb.save(XLSX)

    nb = sum(len(x['miss']) for x in rows_todo)
    print('■ 파일 모음 : %s' % COPYTO)
    print('   PDF %d개 (회차 %d개가 이 파일들을 씁니다)' % (len(copied), len(rows_todo)))
    print('■ 엑셀      : %s' % XLSX)
    print('   정답 입력   %d줄 · 치실 빈칸 %d개' % (len(rows_todo), nb))
    print('   구해야 할 목록 %d줄 (문항 %d개)'
          % (len(rows_need), sum(len(x['need']) for x in rows_need)))


if __name__ == '__main__':
    main()
