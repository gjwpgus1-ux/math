# -*- coding: utf-8 -*-
"""«정답 입력표.xlsx» 에 손으로 채워 넣은 정답을 거두어 온다.

   · 노란 빈칸에 새로 적힌 값만 가져온다 (회색 = 이미 넣어 둔 것은 건드리지 않는다)
   · 결과는 _빌드도구\\정답_엑셀입력.json 에 시험 이름별로 쌓인다
   · 그다음 build_answer.py 를 돌리면 검색기에 들어간다
채운 만큼만 가져오므로 몇 줄만 채우고 돌려도 된다.
"""
import sys, os, re, json, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.dirname(HERE)
DBROOT = os.path.abspath(os.path.join(APP, '..', '데이터베이스'))
XLSX = os.path.join(DBROOT, '정답 입력표.xlsx')
STORE = os.path.join(HERE, '정답_엑셀입력.json')


def main():
    src = open(os.path.join(APP, 'data', 'index.js'), encoding='utf-8').read()
    Q = json.loads(src[src.index('{'):src.rindex('}') + 1].rstrip(';'))
    qno = collections.defaultdict(set)
    for r in Q['items']:
        qno[Q['exams'][r[0]]['n']].add(r[1])
    names = {e['n'] for e in Q['exams']}

    ws = openpyxl.load_workbook(XLSX)['정답 입력']

    old = {}
    if os.path.exists(STORE):
        old = {k: v for k, v in json.load(open(STORE, encoding='utf-8')).items()
               if not k.startswith('_')}

    got = collections.defaultdict(dict)
    bad = []
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 1).value
        if n not in names:
            bad.append((r, n, '모르는 시험 이름')); continue
        for q in range(1, 31):
            v = ws.cell(r, 7 + q).value
            if v is None or v == '-':
                continue
            if isinstance(v, str):
                v = v.strip()
                if not v:
                    continue
                if not re.fullmatch(r'\d{1,4}', v):
                    bad.append((r, n, '%d번 «%s» 는 숫자가 아닙니다' % (q, v))); continue
                v = int(v)
            if not isinstance(v, int) or v < 1:
                bad.append((r, n, '%d번 값이 이상합니다 (%r)' % (q, v))); continue
            if q not in qno[n]:
                bad.append((r, n, '%d번은 그 시험에 없는 번호입니다' % q)); continue
            got[n][str(q)] = v

    # 기존에 거둔 것과 합친다 (엑셀 쪽이 최신)
    out = {k: dict(v) for k, v in old.items()}
    added = changed = 0
    for n, d in got.items():
        cur = out.setdefault(n, {})
        for q, v in d.items():
            if q not in cur:
                cur[q] = v; added += 1
            elif cur[q] != v:
                cur[q] = v; changed += 1

    out = {k: {q: v for q, v in sorted(d.items(), key=lambda t: int(t[0]))}
           for k, d in sorted(out.items()) if d}
    body = {'_설명': '«정답 입력표.xlsx» 에서 거두어 온 정답. read_answer_form.py 가 만든다.'}
    body.update(out)
    json.dump(body, open(STORE, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

    tot = sum(len(d) for d in out.values())
    print('엑셀에서 거둔 정답 : 새로 %d개 · 고쳐 씀 %d개' % (added, changed))
    print('쌓인 것 모두       : 시험 %d회차 · 정답 %d개' % (len(out), tot))
    print('저장               : %s' % STORE)
    if bad:
        print('\n!! 그냥 넘어간 칸 %d개' % len(bad))
        for r, n, m in bad[:12]:
            print('   %d번째 줄 %-22s %s' % (r, n or '', m))
    print('\n이제 build_answer.py 를 돌리면 검색기에 들어갑니다.')


if __name__ == '__main__':
    main()
