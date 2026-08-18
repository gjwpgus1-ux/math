# -*- coding: utf-8 -*-
"""해설지 첫 쪽의 «빠른 정답» 표만 오려 낸다.

  ① 글자를 읽어 «1③2⑤…» 가 있는 줄의 띠를 찾는다 (대부분 여기서 끝난다)
  ② 글꼴 때문에 글자를 못 읽는 옛 파일은, 그림에서 «나란한 가로줄 여섯 개»
     — 정답표의 칸 선 — 를 찾아 그 상자를 오려 낸다
  ③ 오려 낸 표를 시험 이름과 함께 한 장씩 얹어 «정답표 모음.pdf» 로 묶는다
"""
import sys, os, re, json, glob, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import numpy as np
import pypdfium2 as pdfium
from PIL import Image, ImageDraw, ImageFont
import layout

HERE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.dirname(HERE)
DBROOT = os.path.abspath(os.path.join(APP, '..', '데이터베이스'))
SRC = os.path.join(DBROOT, '정답 입력할 파일')
OUT = os.path.join(SRC, '정답표 그림')
BOOK = os.path.join(DBROOT, '정답표 모음.pdf')
XLSX = os.path.join(DBROOT, '정답 입력표.xlsx')

RUN = re.compile(r'(?:\d{1,2}[①②③④⑤]){2,}')
SCALE = 6.0
KFONT = '/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc'


def kfont(size):
    for p in (KFONT,
              '/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc',
              '/usr/share/fonts/truetype/noto/NotoSansCJKkr-Bold.otf'):
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    return ImageFont.load_default()


def band_by_text(pg, lo=1, hi=21, after=4.4):
    """글자로 표의 위·아래·좌·우를 잡는다 → (x0,y0,x1,y1) pt 또는 None

    · 가로: «1③2⑤…» 를 이룬 글자만 따라간다. 줄 전체를 쓰면 옆 단 글까지 물린다.
    · 세로: 찾은 줄은 대개 넷(1~20번)뿐이고 아래 단답형 두 줄은 못 찾으므로,
            줄 간격을 재어 두 줄 반만큼 더 내려간다."""
    ch, W, H = layout.get_chars(pg)
    if not ch:
        return None, None, None
    OK = set('①②③④⑤0123456789')
    rows = []
    for l in layout.make_lines(ch):
        cs = sorted(l, key=lambda c: c[1])
        m = RUN.search(''.join(x[0] for x in cs))
        if not m:
            continue
        first = int(re.match(r'(\d{1,2})', m.group()).group(1))
        if not (lo <= first <= hi):
            continue
        top = min(c[2] for c in l)
        # 표가 시작하는 글자에서 출발해 숫자·동그라미가 이어지는 데까지.
        # 줄 맨 앞에서 출발하면 옆 단 글을 물고 들어온다.
        i0 = m.start()
        x0 = cs[i0][1]; x1 = cs[i0][3]
        for c in cs[i0:]:
            if c[0] in OK and c[1] - x1 < 26:
                x1 = max(x1, c[3])
            elif c[1] - x1 >= 26:
                break
        rows.append((top, x0, x1))
    if not rows:
        return None, W, H
    tops = sorted(r[0] for r in rows)
    gaps = [b - a for a, b in zip(tops, tops[1:]) if 4 < b - a < 60]
    gap = sorted(gaps)[len(gaps) // 2] if gaps else 14.0
    y0 = tops[0] - gap * 1.7
    y1 = tops[-1] + gap * after               # 아래 단답형 줄까지
    x0 = min(r[1] for r in rows) - 10
    x1 = max(r[2] for r in rows) + 12
    return (max(0, x0), max(0, y0), min(W, x1), min(H, y1)), W, H


def band_by_lines(pg, W, H):
    """그림에서 «고르게 벌어진 가로줄 대여섯 개» 를 찾아 표 상자를 잡는다"""
    im = np.array(pg.render(scale=2).to_pil().convert('L'))
    Hh, Ww = im.shape
    dark = im < 128
    half = dark[:, :int(Ww * 0.60)]
    wide = [y for y in range(Hh) if half[y].sum() > Ww * 0.16]
    if len(wide) < 4:
        return None
    # 이어진 줄은 하나로
    lines, run = [], [wide[0]]
    for y in wide[1:]:
        if y - run[-1] <= 2:
            run.append(y)
        else:
            lines.append(sum(run) // len(run)); run = [y]
    lines.append(sum(run) // len(run))
    # 간격이 고른 무리 (정답표는 여섯 줄이 같은 간격)
    best = None
    for i in range(len(lines)):
        grp = [lines[i]]
        for j in range(i + 1, len(lines)):
            d = lines[j] - grp[-1]
            if d < 8:
                continue
            if len(grp) == 1 or abs(d - (grp[1] - grp[0])) <= max(4, (grp[1] - grp[0]) * 0.22):
                grp.append(lines[j])
            else:
                break
        if len(grp) >= 5 and (best is None or len(grp) > len(best)):
            best = grp
    if not best:
        return None
    y0, y1 = best[0] - 10, best[-1] + 10
    cols = np.where(dark[y0:y1].sum(0) > (y1 - y0) * 0.55)[0]
    if len(cols) < 2:
        return None
    x0, x1 = cols.min() - 10, cols.max() + 10
    k = 2.0                                        # 위에서 scale=2 로 그렸다
    return (x0 / k, y0 / k, x1 / k, y1 / k)


SUBJ = {}
SECT = re.compile(r'\[\s*(확률과\s*통계|미적분|기하)\s*\]')


def cut(pg, box):
    x0, y0, x1, y1 = box
    im = pg.render(scale=SCALE).to_pil()
    return im.crop((int(x0 * SCALE), int(y0 * SCALE), int(x1 * SCALE), int(y1 * SCALE)))


def crop_one(path):
    """→ ([(딱지, 그림)], 찾은 방법).  고3은 선택과목 표가 뒷쪽에 따로 있다."""
    # 파일 이름에 과목이 박혀 있으면 (2021년 7월처럼 과목별로 나온 해)  그것을 쓴다
    base = os.path.basename(path)
    fixed = None
    for k, v in (('확통', '확률과 통계'), ('확률과 통계', '확률과 통계'),
                 ('미적', '미적분'), ('기하', '기하')):
        if k in base:
            fixed = v
            break
    d = pdfium.PdfDocument(path)
    out, how = [], '글자'
    try:
        pg = d[0]
        try:
            box, W, H = band_by_text(pg)
            if box is None:
                box = band_by_lines(pg, W, H); how = '줄무늬'
            if box is None:
                box = (0, H * 0.06, W * 0.46, H * 0.40); how = '어림'
            out.append(('', cut(pg, box)))
        finally:
            pg.close()

        # 뒷쪽의 선택과목 표 (23번부터 시작하는 것)
        order = ['확률과 통계', '미적분', '기하']
        k = 0
        for pi in range(1, len(d)):
            pg = d[pi]
            try:
                ch, W, H = layout.get_chars(pg)
                if not ch:
                    continue
                lab, labbox = None, None
                for l in layout.make_lines(ch):
                    cs = sorted(l, key=lambda c: c[1])
                    m = SECT.search(''.join(x[0] for x in cs))
                    if m:
                        lab = m.group(1).replace(' ', '')
                        # 딱지를 이룬 글자만 — 줄 전체를 쓰면 옆 단까지 물린다
                        pc = cs[m.start():m.end()]
                        labbox = (min(c[1] for c in pc), min(c[2] for c in pc),
                                  max(c[3] for c in pc), max(c[4] for c in pc))
                        break
                # 23번부터 시작하는 표만 — 그 쪽의 본문 글은 걸리지 않는다
                sub, W2, H2 = band_by_text(pg, lo=23, hi=23, after=2.9)
                if sub is None:
                    # 글자로 못 읽는 쪽 — 과목 딱지 바로 아래를 오려 낸다.
                    # 딱지는 표 한가운데 위에 놓이므로 좌우로 반 단씩 잡는다.
                    if not labbox:
                        continue
                    cx = (labbox[0] + labbox[2]) / 2
                    sub = (max(0, cx - 118), labbox[1] - 4,
                           min(W, cx + 118), min(H, labbox[3] + 62))
                name = fixed or lab or (order[k] if k < len(order) else '선택')
                k += 1
                out.append((name, cut(pg, sub)))
            finally:
                pg.close()
    finally:
        d.close()
    return out, how


def main():
    os.makedirs(OUT, exist_ok=True)
    for old in glob.glob(os.path.join(OUT, '*.png')):
        os.remove(old)

    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['정답 입력']
    global SUBJ
    SUBJ = {ws.cell(r, 1).value: (ws.cell(r, 5).value or '')
            for r in range(2, ws.max_row + 1)}
    order, rowsof = [], collections.defaultdict(list)
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 6).value
        if f not in rowsof:
            order.append(f)
        rowsof[f].append((r, ws.cell(r, 1).value))

    pages, how_cnt, gaps = [], collections.Counter(), []
    MW = 1500
    for i, f in enumerate(order, 1):
        parts, how = crop_one(os.path.join(SRC, f))
        how_cnt[how] += 1

        # 이 파일이 맡은 시험 가운데 표를 못 찾은 것이 있는지
        got = set()
        for lab, _ in parts:
            got.add(lab.replace(' ', '') if lab else '공통')
        for _, nm in rowsof[f]:
            s = SUBJ.get(nm, '')                       # 그 시험의 과목
            key = {'확통': '확률과통계', '확률과 통계': '확률과통계',
                   '미적': '미적분', '미적분': '미적분',
                   '기하': '기하'}.get(s, '공통')       # 가형·나형·공통·빈칸은 다 공통
            if key not in got:
                gaps.append((f, nm))
        for j, (lab, im) in enumerate(parts):
            im.save(os.path.join(OUT, f[:-4] + (('_' + lab) if lab else '') + '.png'),
                    optimize=True)

        # 한 장으로 꾸미기 — 위에 시험 이름, 아래에 표(들)
        bodies = []
        for lab, im in parts:
            s = min(1.0, MW / im.width)
            bodies.append((lab, im.resize((int(im.width * s), int(im.height * s)))))
        pad, top = 40, 150
        gap = 26
        bh = sum(b.height for _, b in bodies) + gap * len(bodies) + \
             sum(34 for lab, _ in bodies if lab)
        page = Image.new('RGB', (max(MW, max(b.width for _, b in bodies)) + pad * 2,
                                 bh + top + pad), 'white')
        dr = ImageDraw.Draw(page)
        names = ' · '.join(n for _, n in rowsof[f])
        dr.text((pad, 26), '%d / %d      %s' % (i, len(order), f), font=kfont(30), fill=(120, 125, 132))
        fs = 38 if len(names) <= 34 else (30 if len(names) <= 46 else 24)
        dr.text((pad, 74 - fs // 3), names, font=kfont(fs), fill=(20, 30, 60))
        dr.text((pad, 118), '엑셀 «정답 입력» 시트 ' +
                ', '.join(str(r) for r, _ in rowsof[f]) + '번째 줄',
                font=kfont(26), fill=(150, 120, 40))
        y = top
        for lab, b in bodies:
            if lab:
                dr.text((pad, y), '[' + lab + ']', font=kfont(30), fill=(30, 79, 160))
                y += 34
            page.paste(b, (pad, y))
            y += b.height + gap
        pages.append(page)

        for r, _ in rowsof[f]:
            ws.cell(r, 38).value = i         # AL열 — 있던 칸은 건드리지 않는다

    pages[0].save(BOOK, save_all=True, append_images=pages[1:], resolution=150)

    # 엑셀에 «정답표 쪽» 칸 덧붙이기 (파일이 열려 있으면 건너뛴다)
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.cell(1, 38).value = '정답표 쪽'
    ws.cell(1, 38).font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws.cell(1, 38).fill = PatternFill('solid', fgColor='1F3864')
    ws.cell(1, 38).alignment = Alignment(horizontal='center', vertical='center')
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 38).font = Font(name='Arial', size=10)
        ws.cell(r, 38).alignment = Alignment(horizontal='center')
    ws.column_dimensions['AL'].width = 10
    try:
        wb.save(XLSX)
        print('엑셀에 «정답표 쪽» 칸을 붙였습니다.')
    except PermissionError:
        print('!! 엑셀이 열려 있어 «정답표 쪽» 칸을 못 붙였습니다. 닫고 다시 돌려 주세요.')
    tot = sum(os.path.getsize(os.path.join(OUT, x)) for x in os.listdir(OUT))
    print('정답표 그림 %d장 · %.1fMB' % (len(order), tot / 1e6))
    print('찾은 방법:', dict(how_cnt))
    if gaps:
        print('\n!! 표를 못 찾은 시험 %d개 — 해설지에 그 과목 정답표가 없거나 못 읽습니다' % len(gaps))
        for f, nm in gaps:
            print('   %-34s %s' % (f[:34], nm))
    print('묶음 파일 : %s (%.1fMB)' % (BOOK, os.path.getsize(BOOK) / 1e6))


if __name__ == '__main__':
    main()
